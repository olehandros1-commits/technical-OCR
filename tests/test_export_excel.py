from pathlib import Path

from dobs.presentation.export.excel import export_workbook


def _result(
    period_start: str,
    period_end: str,
    beginning: float,
    ending: float,
    deposits: float,
    withdrawals: float,
) -> dict:
    return {
        "account": {
            "bank": "Test Bank",
            "account_last4": "4664",
            "period": {"start": period_start, "end": period_end},
        },
        "summary": {
            "beginning_balance": beginning,
            "ending_balance": ending,
            "deposits_total": deposits,
            "deposits_count": 2,
            "withdrawals_total": withdrawals,
            "withdrawals_count": 1,
            "currency": "USD",
        },
        "transactions": [
            {
                "date": period_start,
                "description": "DEP A",
                "deposit": deposits / 2,
                "withdrawal": None,
                "category": "DEPOSIT_REMOTE",
                "vendor": None,
                "confidence": 0.95,
            },
            {
                "date": period_start,
                "description": "DEP B",
                "deposit": deposits / 2,
                "withdrawal": None,
                "category": "VENDOR_PAYMENT",
                "vendor": "Acme",
                "confidence": 0.88,
            },
            {
                "date": period_end,
                "description": "PAY",
                "deposit": None,
                "withdrawal": withdrawals,
                "category": "ACH_PAYABLE",
                "vendor": None,
                "confidence": 0.40,
            },
        ],
        "_reconciliation": {
            "ok": True,
            "deposits_sum": deposits,
            "withdrawals_sum": withdrawals,
            "deposits_count_actual": 2,
            "withdrawals_count_actual": 1,
            "deposits_total_delta": 0,
            "withdrawals_total_delta": 0,
            "deposits_count_delta": 0,
            "withdrawals_count_delta": 0,
            "balance_equation_delta": 0,
            "issues": [],
        },
        "_anomalies": [],
        "_skipped_rows": [],
    }


def test_workbook_has_all_sheets(tmp_path: Path):
    results = [
        _result("2025-01-01", "2025-01-31", 1000, 1300, 500, 200),
        _result("2025-02-01", "2025-02-28", 1300, 1450, 400, 250),
    ]
    out = tmp_path / "wb.xlsx"
    path = export_workbook(results, out)
    assert path.exists()

    from openpyxl import load_workbook

    wb = load_workbook(path)
    expected = {"Cover", "Summary", "Transactions", "Categories", "Anomalies", "Reconciliation"}
    assert expected.issubset(set(wb.sheetnames))

    # SUMIF formula in Categories
    cats = wb["Categories"]
    has_sumif = any(
        isinstance(c.value, str) and c.value.startswith("=SUMIF")
        for row in cats.iter_rows()
        for c in row
    )
    assert has_sumif, "Categories sheet should contain SUMIF formulas"

    # Continuity formula on Cover
    cover = wb["Cover"]
    has_continuity = any(
        isinstance(c.value, str) and "DRIFT" in (c.value or "")
        for row in cover.iter_rows()
        for c in row
    )
    assert has_continuity, "Cover sheet should contain continuity audit formula"
