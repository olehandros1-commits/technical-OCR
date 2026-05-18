"""Stage 1: Ingest text from any plausible bank-statement input.

Supported inputs (auto-detected by magic bytes / extension):
  * PDF (text-PDF, scanned PDF, encrypted PDF)
  * Plain text (.txt)
  * Image (PNG, JPG, JPEG, TIFF, BMP) -- single page, e.g. phone photo
  * Excel (.xlsx, .xls) -- some banks ship the activity log as a workbook
  * HTML (.html, .htm) -- online banking export

Failure modes handled with explicit errors (no silent bad data):
  * Encrypted PDF without password -> EncryptedDocumentError
  * Truly corrupt input -> CorruptDocumentError
  * Empty / zero-byte -> EmptyDocumentError
  * Unknown format -> UnsupportedFormatError

OCR selection:
  ocr_mode="auto":      text-extract first; if < 200 chars, fall back to
                        Tesseract; if a vision-capable backend is supplied,
                        prefer it on scanned PDFs.
  ocr_mode="tesseract": force Tesseract.
  ocr_mode="vision":    force the vision LLM. Requires a backend.
  ocr_mode="skip":      never OCR; use embedded text or fail.

Speed:
  * Tesseract OCR is **parallelised** across CPU cores via a
    ProcessPoolExecutor (one process per page). 99-page scans drop from
    ~10 min to ~30 s on an 8-core box.
  * OCR results are **cached by file hash** in SQLite -- the same PDF
    uploaded twice returns the text instantly the second time.
  * Each page OCR completion emits a `log_event` so the SSE stream
    shows live progress ('OCR page 12/99').
"""
from __future__ import annotations
import hashlib
import logging
import os
import re
import sqlite3
import threading
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Callable, Optional

log = logging.getLogger(__name__)


# ---- OCR file cache (sqlite) -------------------------------------------

_OCR_CACHE_PATH = Path(os.getenv("OCR_CACHE_DB", "out/ocr_cache.db"))
_OCR_LOCK = threading.Lock()


def _ocr_cache_get(file_hash: str, method: str) -> str | None:
    """Cache lookup is keyed on (file_hash, method): Tesseract and vision
    produce different text from the same PDF, so a vision request must
    not return a stale Tesseract result."""
    if not _OCR_CACHE_PATH.exists():
        return None
    try:
        with _OCR_LOCK, sqlite3.connect(_OCR_CACHE_PATH) as conn:
            row = conn.execute(
                "SELECT text FROM ocr_cache WHERE file_hash = ? AND method = ?",
                (file_hash, method),
            ).fetchone()
            return row[0] if row else None
    except sqlite3.Error:
        return None


def _ocr_cache_put(file_hash: str, text: str, method: str) -> None:
    _OCR_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with _OCR_LOCK, sqlite3.connect(_OCR_CACHE_PATH) as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS ocr_cache (
                file_hash  TEXT NOT NULL,
                method     TEXT NOT NULL,
                text       TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (file_hash, method)
            );
        """)
        conn.execute(
            "INSERT OR REPLACE INTO ocr_cache (file_hash, method, text) VALUES (?, ?, ?)",
            (file_hash, method, text),
        )


def _file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ---- Errors ---------------------------------------------------------------

class IngestError(Exception):
    """Base class for all ingest failures (user-fixable)."""


class EmptyDocumentError(IngestError):
    """Zero-byte or near-zero file."""


class CorruptDocumentError(IngestError):
    """File header is right but parsing failed."""


class EncryptedDocumentError(IngestError):
    """PDF is password-protected and no password was given."""


class UnsupportedFormatError(IngestError):
    """File extension / magic bytes don't match anything we understand."""


# ---- Noise cleanup --------------------------------------------------------

# Common noise from Notion / Azure DI / OCR engines.
_NOTION_HEADER_RE = re.compile(r"\d{2}\.\d{2}\.\d{4},\s*\d{2}:\d{2}\s*", re.MULTILINE)
_NOTION_URL_RE = re.compile(
    r"https?://file\.notion\.so/[^\s]+|file\.notion\.so/[^\s]+", re.IGNORECASE
)
_PAGE_COUNTER_RE = re.compile(r"^\s*\d+/\d+\s*$", re.MULTILINE)
_AZURE_MARKER_RE = re.compile(r":(?:unselected|selected):")


def clean_text(raw: str) -> str:
    text = _NOTION_HEADER_RE.sub("", raw)
    text = _NOTION_URL_RE.sub("", text)
    text = _PAGE_COUNTER_RE.sub("", text)
    text = _AZURE_MARKER_RE.sub("", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


# ---- Format detection -----------------------------------------------------

def _detect_kind(path: Path) -> str:
    """Returns one of: pdf, image, xlsx, html, text, unknown.

    Uses magic bytes first (more robust than extension), with extension as a
    fallback for ambiguous cases.
    """
    try:
        head = path.open("rb").read(8)
    except OSError as e:
        raise IngestError(f"Cannot read file: {e}") from e

    if not head:
        raise EmptyDocumentError(f"File is empty: {path}")

    if head.startswith(b"%PDF-"):
        return "pdf"
    if head.startswith(b"\x89PNG\r\n\x1a\n") or head[:3] in (b"\xff\xd8\xff", b"GIF"):
        return "image"
    if head[:4] == b"PK\x03\x04" and path.suffix.lower() in (".xlsx", ".xlsm"):
        return "xlsx"
    if path.suffix.lower() == ".xls" and head[:4] in (b"\xd0\xcf\x11\xe0",):
        return "xlsx"  # old XLS handled by openpyxl/xlrd; we'll degrade if needed
    if path.suffix.lower() in (".html", ".htm"):
        return "html"
    if head[:5].lower().lstrip() in (b"<!doc", b"<html"):
        return "html"
    if path.suffix.lower() == ".txt":
        return "text"
    # Try heuristic: looks like text?
    try:
        head.decode("utf-8")
        if path.suffix.lower() in (".tif", ".tiff", ".bmp"):
            return "image"
        return "text"
    except UnicodeDecodeError:
        return "unknown"


# ---- Per-format loaders --------------------------------------------------

def from_text_file(path: str | Path) -> str:
    text = Path(path).read_text(encoding="utf-8", errors="replace")
    if not text.strip():
        raise EmptyDocumentError(f"Text file has no content: {path}")
    return clean_text(text)


def from_pdf_text(path: str | Path) -> str:
    """Extract embedded text via pdfplumber. Raises on encrypted/corrupt PDFs.

    When pdfplumber can recognise tabular regions (text-PDFs from Chase,
    BofA, etc.), they are emitted as markdown pipe-tables FIRST, then
    the plain text follows. The downstream regex pre-parser handles both
    representations, but markdown tables let the LLM categorise rows
    with far fewer ambiguities than flowing text.
    """
    import pdfplumber

    try:
        with pdfplumber.open(str(path)) as pdf:
            if getattr(pdf.metadata, "Encrypted", False) or _is_encrypted_pdf(path):
                raise EncryptedDocumentError(
                    f"PDF is encrypted: {path}. Decrypt before passing in."
                )
            parts: list[str] = []
            for page in pdf.pages:
                # Try table extraction first -- bank statements that are
                # native PDFs (not scans) usually have clean column
                # geometry pdfplumber can detect.
                try:
                    tables = page.extract_tables() or []
                except Exception:
                    tables = []
                for tbl in tables:
                    md = _table_to_markdown(tbl)
                    if md:
                        parts.append(md)
                t = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
                parts.append(t)
    except EncryptedDocumentError:
        raise
    except Exception as e:
        # pdfplumber raises many flavours of error on corrupt PDFs.
        if _is_encrypted_pdf(path):
            raise EncryptedDocumentError(str(e)) from e
        raise CorruptDocumentError(f"PDF parse failed: {e}") from e
    return clean_text("\n".join(parts))


def _table_to_markdown(table: list[list[object]]) -> str:
    """Render a pdfplumber table as a pipe-style markdown block.

    Empty cells become blanks; rows that are entirely empty are skipped.
    The first row is treated as the header. Cells longer than 80 chars
    are truncated -- bank descriptions are short by convention.
    """
    if not table or len(table) < 2:
        return ""
    cleaned: list[list[str]] = []
    for row in table:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        cleaned.append([
            (str(c) if c is not None else "").strip().replace("|", "/")[:80]
            for c in row
        ])
    if len(cleaned) < 2:
        return ""
    header = cleaned[0]
    sep = ["---"] * len(header)
    lines = ["| " + " | ".join(header) + " |",
             "| " + " | ".join(sep) + " |"]
    for row in cleaned[1:]:
        # Pad short rows with empty cells so pipe count matches.
        if len(row) < len(header):
            row = row + [""] * (len(header) - len(row))
        lines.append("| " + " | ".join(row[: len(header)]) + " |")
    return "\n".join(lines)


def _is_encrypted_pdf(path: str | Path) -> bool:
    """Conservative check via pikepdf -- handles owner-only encryption too."""
    try:
        import pikepdf
    except ImportError:
        return False
    try:
        pikepdf.open(str(path)).close()
        return False
    except pikepdf.PasswordError:
        return True
    except Exception:
        return False


def _ocr_one_image(img, lang: str, config: str) -> str:
    """OCR one PIL image and return the text. Runs in a worker thread.

    Tesseract releases the GIL during its C work, so a ThreadPoolExecutor
    gives us real parallelism without the ProcessPool start-up + pickling
    overhead that was previously eating 2+ minutes of wall time on a
    99-page PDF inside a Docker container.
    """
    import pytesseract
    return pytesseract.image_to_string(img, lang=lang, config=config)


def from_pdf_ocr(
    path: str | Path,
    lang: str = "eng",
    *,
    workers: int | None = None,
    log_event: Optional[Callable[[str, dict], None]] = None,
) -> str:
    """Tesseract OCR over rasterised pages.

    Parallelises across CPU cores via a ThreadPoolExecutor. Tesseract
    releases the GIL during its C-side OCR work, so threads give real
    parallelism and we avoid ProcessPool start-up + pickling overhead
    (~2 min on a 99-page document inside a Docker container).

    Streaming render: pdf2image now renders pages one at a time, and we
    submit each page to OCR as soon as it's rendered, so OCR work
    overlaps PDF rasterisation.

    Tesseract config:
      * --oem 1 -> LSTM-only engine (skips slow legacy fallback)
      * --psm 4 -> assume a single column of text (right for tabular
        bank statements; ~3-5x faster than the default auto-segmentation)
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    try:
        from pdf2image import convert_from_path
        import pytesseract  # noqa: F401
    except ImportError as exc:
        raise IngestError(
            "Tesseract OCR fallback requires `pdf2image` + `pytesseract` + "
            "Tesseract binary. Install with: pip install bank-statement-extractor[ocr]"
        ) from exc

    dpi = int(os.getenv("OCR_DPI", "120"))
    config = os.getenv("OCR_TESSERACT_CONFIG", "--oem 1 --psm 4")
    if log_event:
        log_event("ocr_render_start", {
            "engine": "tesseract", "dpi": dpi, "config": config,
        })
    t0 = time.time()
    try:
        # Get total page count cheaply via pdfinfo (poppler).
        from pdf2image.pdf2image import pdfinfo_from_path
        info = pdfinfo_from_path(str(path))
        n_pages = int(info.get("Pages", 0))
    except Exception:
        n_pages = 0

    n_workers = workers or max(1, min(int(os.getenv("OCR_WORKERS", "8")), n_pages or 8))
    if log_event:
        log_event("ocr_pool_start", {
            "pages": n_pages or "?", "workers": n_workers,
        })

    results: dict[int, str] = {}
    completed = 0
    pool = ThreadPoolExecutor(max_workers=n_workers)
    futures: dict = {}
    t_ocr = time.time()

    try:
        # Render and dispatch one page at a time so OCR runs in parallel
        # with rendering. pdf2image lets us pass first_page/last_page.
        for page_idx in range(1, (n_pages or 1) + 1):
            try:
                imgs = convert_from_path(
                    str(path), dpi=dpi,
                    first_page=page_idx, last_page=page_idx,
                )
            except Exception as e:
                if page_idx == 1:
                    raise CorruptDocumentError(f"PDF rasterisation failed: {e}") from e
                break  # ran past the end (n_pages was wrong / 0)
            if not imgs:
                break
            fut = pool.submit(_ocr_one_image, imgs[0], lang, config)
            futures[fut] = page_idx - 1
            # Drain any already-completed futures opportunistically.
            for done in list(as_completed(futures, timeout=0)) if False else []:
                # `timeout=0` is undefined in cf; we just check on every
                # render iteration in the explicit drain below.
                pass

        if log_event:
            log_event("ocr_render_done", {
                "pages": len(futures),
                "elapsed_s": round(time.time() - t0, 2),
            })

        # Now wait for all the per-page OCR futures.
        for fut in as_completed(futures):
            i = futures[fut]
            results[i] = fut.result()
            completed += 1
            if log_event and (completed % 5 == 0 or completed == len(futures)):
                log_event("ocr_page", {
                    "done": completed, "of": len(futures),
                    "elapsed_s": round(time.time() - t_ocr, 1),
                })
    finally:
        pool.shutdown(wait=True)

    if log_event:
        log_event("ocr_pool_done", {
            "pages": len(results),
            "elapsed_s": round(time.time() - t_ocr, 2),
        })
    return clean_text("\n".join(results[i] for i in sorted(results)))


def from_image(path: str | Path, lang: str = "eng") -> str:
    """Single image via Tesseract (typical phone-photo case)."""
    try:
        import pytesseract
        from PIL import Image
    except ImportError as exc:
        raise IngestError(
            "Image ingest requires `pytesseract` + `Pillow` + Tesseract binary."
        ) from exc
    try:
        img = Image.open(path)
    except Exception as e:
        raise CorruptDocumentError(f"Image load failed: {e}") from e
    return clean_text(pytesseract.image_to_string(img, lang=lang))


def from_xlsx(path: str | Path) -> str:
    """Flatten an Excel workbook into a single text blob the LLM can read."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise IngestError("Excel ingest requires `openpyxl`.") from exc
    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:
        raise CorruptDocumentError(f"Workbook load failed: {e}") from e
    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f"=== Sheet: {ws.title} ===")
        for row in ws.iter_rows(values_only=True):
            line = " | ".join("" if v is None else str(v) for v in row)
            parts.append(line)
    return clean_text("\n".join(parts))


def from_html(path: str | Path) -> str:
    """Strip tags, keep text."""
    raw = Path(path).read_text(encoding="utf-8", errors="replace")
    # Cheap tag strip; bank online-banking exports are usually simple HTML.
    txt = re.sub(r"<[^>]+>", " ", raw)
    txt = re.sub(r"&nbsp;", " ", txt)
    txt = re.sub(r"&amp;", "&", txt)
    return clean_text(txt)


def from_pdf_vision(path: str | Path, backend) -> str:
    from extractor.ingest_vision import from_pdf_vision as _impl
    return clean_text(_impl(path, backend))


# ---- Public entry point --------------------------------------------------

def ingest(
    pdf_path: str | Path,
    txt_path: str | Path | None = None,
    *,
    ocr_mode: str = "auto",
    backend: Optional[object] = None,
    log_event: Optional[Callable[[str, dict], None]] = None,
) -> str:
    """Pluggable ingest. Returns cleaned text for the whole document.

    See module docstring for ocr_mode semantics. The first positional arg is
    historically called ``pdf_path`` for spec-compatibility, but any
    supported format (image, xlsx, html, text) is accepted.

    Raises ``IngestError`` (or subclass) on bad input so the caller can show
    the user a clean error instead of a stack trace.

    `log_event` is forwarded to the OCR path so per-page progress shows up
    in the SSE stream.
    """
    if txt_path:
        return from_text_file(txt_path)

    main = Path(pdf_path)
    if not main.exists():
        raise IngestError(f"File not found: {main}")
    if main.stat().st_size == 0:
        raise EmptyDocumentError(f"File is empty: {main}")

    kind = _detect_kind(main)
    log.info("ingest: %s detected as %s", main.name, kind)

    if kind == "pdf":
        if ocr_mode == "vision":
            if backend is None:
                raise ValueError("ocr_mode='vision' requires a backend")
            return _cached_ocr(main, "vision", log_event,
                               lambda: from_pdf_vision(main, backend))
        if ocr_mode == "tesseract":
            return _cached_ocr(main, "tesseract", log_event,
                               lambda: from_pdf_ocr(main, log_event=log_event))
        # auto / skip
        text = from_pdf_text(main)
        if ocr_mode == "skip" or len(text.strip()) >= 200:
            return text
        # Auto fallback for scanned PDFs.
        if backend is not None and backend.supports_vision():
            log.info("ingest: text-PDF empty, falling back to vision LLM")
            return _cached_ocr(main, "vision", log_event,
                               lambda: from_pdf_vision(main, backend))
        log.info("ingest: text-PDF empty, falling back to Tesseract")
        return _cached_ocr(main, "tesseract", log_event,
                           lambda: from_pdf_ocr(main, log_event=log_event))

    if kind == "image":
        return from_image(main)
    if kind == "xlsx":
        return from_xlsx(main)
    if kind == "html":
        return from_html(main)
    if kind == "text":
        return from_text_file(main)

    raise UnsupportedFormatError(
        f"Cannot determine how to read {main.name}. "
        "Supported: PDF, image, xlsx, html, txt."
    )


def _cached_ocr(
    path: Path,
    method: str,
    log_event: Optional[Callable[[str, dict], None]],
    do_ocr: Callable[[], str],
) -> str:
    """Wrap any OCR path with (file_hash, method) caching so repeats are instant."""
    h = _file_hash(path)
    cached = _ocr_cache_get(h, method)
    if cached is not None:
        if log_event:
            log_event("ocr_cache_hit", {
                "file_hash": h[:16], "chars": len(cached), "method": method,
            })
        return cached
    if log_event:
        log_event("ocr_cache_miss", {"file_hash": h[:16], "method": method})
    text = do_ocr()
    try:
        _ocr_cache_put(h, text, method)
    except Exception as e:
        log.warning("ocr_cache_put failed: %s", e)
    return text
