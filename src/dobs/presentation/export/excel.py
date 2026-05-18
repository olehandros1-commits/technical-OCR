from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="F9FAFB", bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=16, color="0F172A")
_OK_FILL = PatternFill("solid", fgColor="DCFCE7")
_BAD_FILL = PatternFill("solid", fgColor="FEE2E2")
_WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
_LOW_CONF_FILL = PatternFill("solid", fgColor="FFF7ED")
_THIN = Side(border_style="thin", color="E5E7EB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


class ExcelPresenter:
    """Builds a multi-sheet Excel workbook from extraction results.
    Data shaping via pandas; openpyxl only for live formulas + formatting."""

    def __init__(self, /, *, results: list[dict]) -> None:
        self._results = results
        self._df_summary = self._build_summary_df()
        self._df_transactions = self._build_transactions_df()
        self._df_anomalies = self._build_anomalies_df()
        self._df_reconciliation = self._build_reconciliation_df()

    # ------------------------------------------------------------------ build

    def _build_summary_df(self) -> pd.DataFrame:
        rows = []
        for i, r in enumerate(self._results, start=1):
            a, s = r["account"], r["summary"]
            recon = r.get("_reconciliation") or {}
            rows.append({
                "#": i,
                "Period start": a["period"]["start"],
                "Period end": a["period"]["end"],
                "Account": a.get("account_last4"),
                "Beginning balance": s["beginning_balance"],
                "Ending balance": s["ending_balance"],
                "Deposits count": s.get("deposits_count"),
                "Deposits total": s["deposits_total"],
                "Withdrawals count": s.get("withdrawals_count"),
                "Withdrawals total": s["withdrawals_total"],
                "Reconciled": "OK" if recon.get("ok") else "MISMATCH",
                "Issues": "; ".join(recon.get("issues", []))[:200],
            })
        return pd.DataFrame(rows)

    def _build_transactions_df(self) -> pd.DataFrame:
        rows = []
        for r in self._results:
            period = r["account"]["period"]["start"]
            acct = r["account"].get("account_last4") or ""
            for t in r["transactions"]:
                rows.append({
                    "Statement": period,
                    "Account": acct,
                    "Date": t["date"],
                    "Description": t["description"],
                    "Category": t.get("category"),
                    "Vendor": t.get("vendor"),
                    "Deposit": t.get("deposit"),
                    "Withdrawal": t.get("withdrawal"),
                    "Confidence": t.get("confidence"),
                })
        return pd.DataFrame(rows)

    def _build_anomalies_df(self) -> pd.DataFrame:
        rows = []
        for r in self._results:
            period = r["account"]["period"]["start"]
            acct = r["account"].get("account_last4") or ""
            for an in r.get("_anomalies", []):
                rows.append({
                    "Statement": period,
                    "Account": acct,
                    "Kind": an["kind"],
                    "Severity": an["severity"],
                    "Tx index": an.get("transaction_index"),
                    "Message": an["message"],
                })
        return pd.DataFrame(rows)

    def _build_reconciliation_df(self) -> pd.DataFrame:
        rows = []
        for r in self._results:
            a, s = r["account"], r["summary"]
            recon = r.get("_reconciliation") or {}
            rows.append({
                "Period": a["period"]["start"],
                "Account": a.get("account_last4"),
                "Decl deposits $": s["deposits_total"],
                "Extracted deposits $": recon.get("deposits_sum"),
                "Decl deposits #": s.get("deposits_count"),
                "Extracted deposits #": recon.get("deposits_count_actual"),
                "Decl withdrawals $": s["withdrawals_total"],
                "Extracted withdrawals $": recon.get("withdrawals_sum"),
                "Decl withdrawals #": s.get("withdrawals_count"),
                "Extracted withdrawals #": recon.get("withdrawals_count_actual"),
                "Balance Δ": recon.get("balance_equation_delta"),
                "ok": recon.get("ok", False),
            })
        return pd.DataFrame(rows)

    # ------------------------------------------------------------------ render

    def render(self, out_path: Path) -> Path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        self._add_cover(wb)
        self._add_summary(wb)
        tx_sheet_name = "Transactions"
        tx_row_range = self._add_transactions(wb, sheet_name=tx_sheet_name)
        self._add_categories(wb, tx_sheet=tx_sheet_name, tx_range=tx_row_range)
        self._add_anomalies(wb)
        self._add_reconciliation(wb)
        wb.save(out_path)
        return out_path

    # ------------------------------------------------------------------ sheets

    def _add_cover(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Cover"
        ws["A1"] = "Bank Statement Extraction Report"
        ws["A1"].font = _TITLE_FONT
        ws.merge_cells("A1:F1")

        ws["A3"] = "Generated by"
        ws["B3"] = "dobs-extractor"
        ws["A4"] = "Statements"
        ws["B4"] = len(self._results)
        ws["A5"] = "Reconciled"
        ws["B5"] = sum(
            1 for r in self._results if (r.get("_reconciliation") or {}).get("ok")
        )
        ws["A6"] = "Total transactions"
        ws["B6"] = len(self._df_transactions)
        ws["A7"] = "Total anomalies"
        ws["B7"] = len(self._df_anomalies)

        ws["A9"] = "Coverage by statement"
        ws["A9"].font = Font(bold=True, size=13)

        cover_cols = [
            "Period", "Bank", "Account", "Currency",
            "Beginning $", "Ending $",
            "Deposits #", "Deposits $",
            "Withdrawals #", "Withdrawals $",
            "Reconciled", "Anomalies",
        ]
        cover_header_row = 11
        self._bold_header(ws, cover_header_row, cover_cols)

        for i, r in enumerate(self._results, start=cover_header_row + 1):
            a, s = r["account"], r["summary"]
            recon = r.get("_reconciliation") or {}
            ws.cell(i, 1, f"{a['period']['start']} -> {a['period']['end']}")
            ws.cell(i, 2, a.get("bank"))
            ws.cell(i, 3, a.get("account_last4"))
            ws.cell(i, 4, s.get("currency") or "USD")
            ws.cell(i, 5, s["beginning_balance"]).number_format = "$#,##0.00"
            ws.cell(i, 6, s["ending_balance"]).number_format = "$#,##0.00"
            ws.cell(i, 7, s.get("deposits_count"))
            ws.cell(i, 8, s["deposits_total"]).number_format = "$#,##0.00"
            ws.cell(i, 9, s.get("withdrawals_count"))
            ws.cell(i, 10, s["withdrawals_total"]).number_format = "$#,##0.00"
            ok = recon.get("ok", False)
            ws.cell(i, 11, "OK" if ok else "MISMATCH").fill = _OK_FILL if ok else _BAD_FILL
            ws.cell(i, 12, len(r.get("_anomalies", [])))

        audit_col = 13
        ws.cell(cover_header_row, audit_col, "Continuity?")
        ws.cell(cover_header_row, audit_col).fill = _HEADER_FILL
        ws.cell(cover_header_row, audit_col).font = _HEADER_FONT
        for i in range(cover_header_row + 2, cover_header_row + 1 + len(self._results)):
            ws.cell(i, audit_col, f'=IF(ROUND(E{i}-F{i-1},2)=0,"OK","DRIFT $"&TEXT(F{i-1}-E{i},"#,##0.00"))')

        self._autosize_columns(ws)

    def _add_summary(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Summary")
        df = self._df_summary

        self._bold_header(ws, 1, list(df.columns))
        for i, row in enumerate(df.itertuples(index=False), start=2):
            for j, val in enumerate(row, start=1):
                ws.cell(i, j, val)

        money_cols = {"Beginning balance": 5, "Ending balance": 6, "Deposits total": 8, "Withdrawals total": 10}
        for col_name, col_idx in money_cols.items():
            for i in range(2, 2 + len(df)):
                ws.cell(i, col_idx).number_format = "$#,##0.00"

        for i in range(2, 2 + len(df)):
            recon_val = ws.cell(i, 11).value
            ws.cell(i, 11).fill = _OK_FILL if recon_val == "OK" else _BAD_FILL

        last_row = 1 + len(df)
        totals_row = last_row + 1
        ws.cell(totals_row, 1, "TOTAL").font = Font(bold=True)
        ws.cell(totals_row, 7, f"=SUM(G2:G{last_row})")
        ws.cell(totals_row, 8, f"=SUM(H2:H{last_row})").number_format = "$#,##0.00"
        ws.cell(totals_row, 9, f"=SUM(I2:I{last_row})")
        ws.cell(totals_row, 10, f"=SUM(J2:J{last_row})").number_format = "$#,##0.00"
        for c in range(1, 13):
            ws.cell(totals_row, c).fill = PatternFill("solid", fgColor="E0E7FF")

        self._autosize_columns(ws)

    def _add_transactions(self, wb: Workbook, *, sheet_name: str) -> tuple[int, int]:
        ws = wb.create_sheet(sheet_name)
        df = self._df_transactions

        # Write columns without the formula-derived "Amount" column — that must be a live formula.
        display_cols = ["Statement", "Account", "Date", "Description", "Category", "Vendor", "Deposit", "Withdrawal", "Confidence"]
        full_header = display_cols[:8] + ["Amount", "Confidence"]
        self._bold_header(ws, 1, full_header)

        first_data = 2
        row = first_data
        for rec in df[display_cols].itertuples(index=False):
            # Statement, Account, Date, Description, Category, Vendor, Deposit, Withdrawal
            for j, val in enumerate(rec[:8], start=1):
                ws.cell(row, j, val)
            # Deposit / Withdrawal number formats
            ws.cell(row, 7).number_format = "$#,##0.00"
            ws.cell(row, 8).number_format = "$#,##0.00"
            # Col 9 = Amount: live formula
            ws.cell(row, 9, f"=IF(ISNUMBER(G{row}),G{row},-IFERROR(H{row},0))").number_format = "$#,##0.00;[Red]-$#,##0.00"
            # Col 10 = Confidence (index 8 in the named tuple)
            ws.cell(row, 10, rec[8])
            row += 1

        last_data = row - 1
        if last_data >= first_data:
            ws.conditional_formatting.add(
                f"A{first_data}:J{last_data}",
                FormulaRule(formula=[f"$J{first_data}<0.5"], fill=_LOW_CONF_FILL, stopIfTrue=False),
            )
            ws.conditional_formatting.add(
                f"J{first_data}:J{last_data}",
                ColorScaleRule(
                    start_type="num", start_value=0, start_color="FCA5A5",
                    mid_type="num", mid_value=0.5, mid_color="FDE68A",
                    end_type="num", end_value=1, end_color="86EFAC",
                ),
            )
        ws.freeze_panes = "A2"
        self._autosize_columns(ws)
        return first_data, last_data

    def _add_categories(self, wb: Workbook, *, tx_sheet: str, tx_range: tuple[int, int]) -> None:
        ws = wb.create_sheet("Categories")
        first, last = tx_range
        self._bold_header(ws, 1, ["Category", "Count (formula)", "Total amount (formula)"])

        cats_sorted = sorted(
            {t.get("category") for r in self._results for t in r["transactions"] if t.get("category")}
        ) or ["(none)"]

        for i, cat in enumerate(cats_sorted, start=2):
            ws.cell(i, 1, cat)
            ws.cell(i, 2, f"=COUNTIF({tx_sheet}!E{first}:E{last},A{i})")
            ws.cell(
                i, 3,
                f"=SUMIF({tx_sheet}!E{first}:E{last},A{i},{tx_sheet}!I{first}:I{last})",
            ).number_format = "$#,##0.00;[Red]-$#,##0.00"

        self._autosize_columns(ws)

    def _add_anomalies(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Anomalies")
        df = self._df_anomalies
        self._bold_header(ws, 1, list(df.columns))

        severity_fill = {"info": _OK_FILL, "warn": _WARN_FILL, "error": _BAD_FILL}
        for i, rec in enumerate(df.itertuples(index=False), start=2):
            for j, val in enumerate(rec, start=1):
                ws.cell(i, j, val)
            ws.cell(i, 4).fill = severity_fill.get(ws.cell(i, 4).value, _OK_FILL)

        ws.freeze_panes = "A2"
        self._autosize_columns(ws)

    def _add_reconciliation(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Reconciliation")
        cols = [
            "Period", "Account",
            "Decl deposits $", "Extracted $", "Δ$",
            "Decl deposits #", "Extracted #", "Δ#",
            "Decl withdrawals $", "Extracted $", "Δ$",
            "Decl withdrawals #", "Extracted #", "Δ#",
            "Balance Δ", "Status",
        ]
        self._bold_header(ws, 1, cols)

        df = self._df_reconciliation
        for i, rec in enumerate(df.itertuples(index=False), start=2):
            # rec fields match _build_reconciliation_df column order
            ws.cell(i, 1, rec[0])   # Period
            ws.cell(i, 2, rec[1])   # Account
            ws.cell(i, 3, rec[2]).number_format = "$#,##0.00"   # Decl deposits $
            ws.cell(i, 4, rec[3]).number_format = "$#,##0.00"   # Extracted deposits $
            ws.cell(i, 5, f"=C{i}-D{i}").number_format = "$#,##0.00;[Red]-$#,##0.00"  # Δ$
            ws.cell(i, 6, rec[4])   # Decl deposits #
            ws.cell(i, 7, rec[5])   # Extracted deposits #
            ws.cell(i, 8, f"=F{i}-G{i}")  # Δ#
            ws.cell(i, 9, rec[6]).number_format = "$#,##0.00"   # Decl withdrawals $
            ws.cell(i, 10, rec[7]).number_format = "$#,##0.00"  # Extracted withdrawals $
            ws.cell(i, 11, f"=I{i}-J{i}").number_format = "$#,##0.00;[Red]-$#,##0.00"  # Δ$
            ws.cell(i, 12, rec[8])  # Decl withdrawals #
            ws.cell(i, 13, rec[9])  # Extracted withdrawals #
            ws.cell(i, 14, f"=L{i}-M{i}")  # Δ#
            ws.cell(i, 15, rec[10]).number_format = "$#,##0.00;[Red]-$#,##0.00"  # Balance Δ
            ws.cell(i, 16, "OK" if rec[11] else "MISMATCH")  # Status (ok bool)

        last_row = 1 + len(df)
        for col_letter in ("E", "H", "K", "N", "O"):
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{last_row}",
                CellIsRule(operator="notEqual", formula=["0"], fill=_BAD_FILL),
            )
        ws.conditional_formatting.add(
            f"P2:P{last_row}",
            CellIsRule(operator="equal", formula=['"OK"'], fill=_OK_FILL),
        )
        ws.conditional_formatting.add(
            f"P2:P{last_row}",
            CellIsRule(operator="equal", formula=['"MISMATCH"'], fill=_BAD_FILL),
        )
        self._autosize_columns(ws)

    # ------------------------------------------------------------------ helpers

    def _bold_header(self, ws, row: int, columns: list[str]) -> None:
        for i, col in enumerate(columns, start=1):
            c = ws.cell(row=row, column=i, value=col)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _BORDER

    def _autosize_columns(self, ws, max_width: int = 60) -> None:
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=0,
            )
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def export_workbook(results: list[dict], out_path: str | Path) -> Path:
    return ExcelPresenter(results=results).render(Path(out_path))
