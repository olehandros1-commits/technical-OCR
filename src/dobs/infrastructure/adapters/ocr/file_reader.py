from __future__ import annotations

import re
from pathlib import Path


class IngestError(Exception):
    pass


class EmptyDocumentError(IngestError):
    pass


class CorruptDocumentError(IngestError):
    pass


class EncryptedDocumentError(IngestError):
    pass


class UnsupportedFormatError(IngestError):
    pass


_NOTION_HEADER_RE = re.compile(r"\d{2}\.\d{2}\.\d{4},\s*\d{2}:\d{2}\s*", re.MULTILINE)
_NOTION_URL_RE = re.compile(
    r"https?://file\.notion\.so/[^\s]+|file\.notion\.so/[^\s]+", re.IGNORECASE
)
_PAGE_COUNTER_RE = re.compile(r"^\s*\d+/\d+\s*$", re.MULTILINE)
_AZURE_MARKER_RE = re.compile(r":(?:unselected|selected):")


def clean_text(raw: str) -> str:
    text = _NOTION_HEADER_RE.sub("", raw)
    text = _NOTION_URL_RE.sub("", text)
    text = _PAGE_COUNTER_RE.sub("", text)
    text = _AZURE_MARKER_RE.sub("", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def detect_kind(path: Path) -> str:
    try:
        head = path.open("rb").read(8)
    except OSError as e:
        raise IngestError(f"Cannot read file: {e}") from e

    if not head:
        raise EmptyDocumentError(f"File is empty: {path}")

    if head.startswith(b"%PDF-"):
        return "pdf"
    if head.startswith(b"\x89PNG\r\n\x1a\n") or head[:3] in (b"\xff\xd8\xff", b"GIF"):
        return "image"
    if head[:4] == b"PK\x03\x04" and path.suffix.lower() in (".xlsx", ".xlsm"):
        return "xlsx"
    if path.suffix.lower() == ".xls" and head[:4] in (b"\xd0\xcf\x11\xe0",):
        return "xlsx"
    if path.suffix.lower() in (".html", ".htm"):
        return "html"
    if head[:5].lower().lstrip() in (b"<!doc", b"<html"):
        return "html"
    if path.suffix.lower() == ".txt":
        return "text"
    try:
        head.decode("utf-8")
        if path.suffix.lower() in (".tif", ".tiff", ".bmp"):
            return "image"
        return "text"
    except UnicodeDecodeError:
        return "unknown"


class FileReader:
    def __init__(self, /) -> None:
        pass

    async def read(self, path: Path) -> str:
        text = path.read_text(encoding="utf-8", errors="replace")
        if not text.strip():
            raise EmptyDocumentError(f"Text file has no content: {path}")
        return clean_text(text)

    def from_xlsx(self, path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise IngestError("Excel ingest requires `openpyxl`.") from exc
        try:
            wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        except Exception as e:
            raise CorruptDocumentError(f"Workbook load failed: {e}") from e
        parts: list[str] = []
        for ws in wb.worksheets:
            parts.append(f"=== Sheet: {ws.title} ===")
            for row in ws.iter_rows(values_only=True):
                line = " | ".join("" if v is None else str(v) for v in row)
                parts.append(line)
        return clean_text("\n".join(parts))

    def from_html(self, path: Path) -> str:
        raw = path.read_text(encoding="utf-8", errors="replace")
        txt = re.sub(r"<[^>]+>", " ", raw)
        txt = re.sub(r"&nbsp;", " ", txt)
        txt = re.sub(r"&amp;", "&", txt)
        return clean_text(txt)

    def from_image(self, path: Path, lang: str = "eng") -> str:
        try:
            import pytesseract
            from PIL import Image
        except ImportError as exc:
            raise IngestError(
                "Image ingest requires `pytesseract` + `Pillow` + Tesseract binary."
            ) from exc
        try:
            img = Image.open(path)
        except Exception as e:
            raise CorruptDocumentError(f"Image load failed: {e}") from e
        return clean_text(pytesseract.image_to_string(img, lang=lang))

    def from_pdf_text(self, path: Path) -> str:
        import pdfplumber

        try:
            with pdfplumber.open(str(path)) as pdf:
                if getattr(pdf.metadata, "Encrypted", False) or self._is_encrypted_pdf(path):
                    raise EncryptedDocumentError(
                        f"PDF is encrypted: {path}. Decrypt before passing in."
                    )
                parts: list[str] = []
                for page in pdf.pages:
                    try:
                        tables = page.extract_tables() or []
                    except Exception:
                        tables = []
                    for tbl in tables:
                        md = self._table_to_markdown(tbl)
                        if md:
                            parts.append(md)
                    t = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
                    parts.append(t)
        except EncryptedDocumentError:
            raise
        except Exception as e:
            if self._is_encrypted_pdf(path):
                raise EncryptedDocumentError(str(e)) from e
            raise CorruptDocumentError(f"PDF parse failed: {e}") from e
        return clean_text("\n".join(parts))

    def _is_encrypted_pdf(self, path: Path) -> bool:
        try:
            import pikepdf
        except ImportError:
            return False
        try:
            pikepdf.open(str(path)).close()
            return False
        except pikepdf.PasswordError:
            return True
        except Exception:
            return False

    def _table_to_markdown(self, table: list[list[str | None]]) -> str:
        if not table or len(table) < 2:
            return ""
        cleaned: list[list[str]] = []
        for row in table:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            cleaned.append(
                [(str(c) if c is not None else "").strip().replace("|", "/")[:80] for c in row]
            )
        if len(cleaned) < 2:
            return ""
        header = cleaned[0]
        sep = ["---"] * len(header)
        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(sep) + " |",
        ]
        for clean_row in cleaned[1:]:
            if len(clean_row) < len(header):
                clean_row = clean_row + [""] * (len(header) - len(clean_row))
            lines.append("| " + " | ".join(clean_row[: len(header)]) + " |")
        return "\n".join(lines)
